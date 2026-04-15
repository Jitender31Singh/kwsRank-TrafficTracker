import math
import os
import time
import io
import requests
import pandas as pd
import logging
from datetime import date, timedelta
from dotenv import load_dotenv
from google.oauth2 import service_account
from googleapiclient.discovery import build
from openpyxl import load_workbook
import re
from newMarketingWIthWriteLogic  import DRIVE_ID, SITE_ID, SITE_URL, GraphAPIClient,get_gsc_service
from newMarketingWIthWriteLogic import LOCAL_FILE, GraphAPIClient, download_file, flatten_sheet_headers, format_sheet_headers, update_page_sheet, upload_file

load_dotenv()

FILE_ID = os.getenv("PAGE_RANKING_FILE_ID")
LOCAL_FILE = "temp_daily.xlsx"

logger = logging.getLogger(__name__)

def filter_trackable_pages(df_page):
    df_page = df_page.copy()

    if "page" not in df_page.columns:
        raise KeyError(f"'page' column not found in df_page. Found: {df_page.columns.tolist()}")

    # Normalize
    df_page["page"] = df_page["page"].astype(str).str.strip()

    # Remove blanks / nan-like values
    df_page = df_page[
        df_page["page"].notna()
        & (df_page["page"] != "")
        & (df_page["page"].str.lower() != "nan")
    ]

    # Remove unwanted URLs:
    # - images/files
    # - wp-content assets
    # - fragment (#)
    # - query params (?)
    df_page = df_page[
        ~df_page["page"].str.lower().str.contains(
            r"\.webp|\.png|\.jpg|\.jpj|wp-content|#|\?",
            regex=True,
            na=False
        )
    ]

    # # Keep only proper web URLs
    # df_page = df_page[
    #     df_page["page"].str.lower().str.startswith(("http://", "https://"))
    # ]

    logger.info("Filtered trackable page rows count: %s", len(df_page))
    # logger.info("Remaining sample URLs: %s", df_page["page"].head(10).tolist())

    return df_page

def build_edit_update_page_df(df_page_sheet, df_gsc_page, end_date):
    # ==========================================
    # 1) Create normalized temp copy for URL matching only
    # ==========================================
    df_page_sheet = df_page_sheet.copy()
    df_temp = df_page_sheet.copy()

    df_temp.columns = (
        df_temp.columns.astype(str)
        .str.strip()
        .str.lower()
        .str.replace(" ", "_")
    )

    # logger.info("df_temp columns: %s", df_temp.columns.tolist())

    # Find URL column safely
    url_col = "urls"
    if url_col not in df_temp.columns:
        raise KeyError(f"Expected column '{url_col}' in Page sheet. Found: {list(df_temp.columns)}")

    # Normalize URLs in page sheet
    df_temp[url_col] = df_temp[url_col].astype(str).str.lower().str.strip()
    df_page_sheet["urls"] = df_temp[url_col]

    # ==========================================
    # 2) Normalize GSC page URLs
    # ==========================================
    df_gsc_page = df_gsc_page.copy()
    df_gsc_page["page"] = df_gsc_page["page"].astype(str).str.lower().str.strip()

    # ==========================================
    # 3) Sort by impressions DESC and remove duplicate pages
    #    (NO aggregation, just keep best row per page)
    # ==========================================
    df_gsc_page = df_gsc_page.sort_values(by="impressions", ascending=False)
    df_gsc_unique = df_gsc_page.drop_duplicates(subset=["page"], keep="first").copy()

    # ==========================================
    # 4) Remove duplicate URLs from existing sheet
    #    Final output should not contain duplicates
    # ==========================================
    df_page_sheet = df_page_sheet.drop_duplicates(subset=["urls"], keep="first").reset_index(drop=True)

    # ==========================================
    # 5) New column names for current date
    # ==========================================
    clicks_col = f"{end_date}_Clicks"
    impressions_col = f"{end_date}_Impressions"
    ctr_col = f"{end_date}_CTR"
    position_col = f"{end_date}_Avg Position"

    # Ensure columns exist in df_page_sheet
    for col in [clicks_col, impressions_col, ctr_col, position_col]:
        if col not in df_page_sheet.columns:
            df_page_sheet[col] = pd.NA

    # ==========================================
    # 6) Update ALL existing URLs from full GSC data (all 25000 rows, deduped by page only)
    # ==========================================
    metrics_lookup = df_gsc_unique.set_index("page")

    df_page_sheet[clicks_col] = df_page_sheet["urls"].map(metrics_lookup["clicks"])
    df_page_sheet[impressions_col] = df_page_sheet["urls"].map(metrics_lookup["impressions"])
    df_page_sheet[ctr_col] = df_page_sheet["urls"].map(metrics_lookup["ctr"])
    df_page_sheet[position_col] = df_page_sheet["urls"].map(metrics_lookup["position"])

    # ==========================================
    # 7) Take TOP 100 URLs by impressions from GSC
    # ==========================================
    top_100_pages = df_gsc_unique.head(100).copy()

    # ==========================================
    # 8) Insert only TOP 100 URLs that are NOT already present in sheet
    # ==========================================
    existing_urls = set(df_page_sheet["urls"].dropna())
    new_pages = top_100_pages[~top_100_pages["page"].isin(existing_urls)].copy()

    # ==========================================
    # 9) Append new rows for missing TOP 100 URLs
    # ==========================================
    if not new_pages.empty:
        new_rows = pd.DataFrame(columns=df_page_sheet.columns)

        new_rows["urls"] = new_pages["page"].values
        new_rows[clicks_col] = new_pages["clicks"].values
        new_rows[impressions_col] = new_pages["impressions"].values
        new_rows[ctr_col] = new_pages["ctr"].values
        new_rows[position_col] = new_pages["position"].values

        df_page_sheet = pd.concat([df_page_sheet, new_rows], ignore_index=True)

    # ==========================================
    # 10) Final safety dedupe (no duplicate URLs in output)
    # ==========================================
    df_page_sheet = df_page_sheet.drop_duplicates(subset=["urls"], keep="first").reset_index(drop=True)

    return df_page_sheet

def fetch_data(service, start_date, end_date, dimension):
    logger.info(f"Fetching GSC data from {start_date} to {end_date}")

    response = service.searchanalytics().query(
        siteUrl=SITE_URL,
        body={
            "startDate": str(start_date),
            "endDate": str(end_date),
            "dimensions": dimension,
            "rowLimit": 25000
        }
    ).execute()

    dim_name = dimension[0]  # "query" OR "page"

    rows = []
    for row in response.get("rows", []):
        rows.append({
            dim_name: row["keys"][0],
            "clicks": row.get("clicks", 0),
            "impressions": row.get("impressions", 0),
            "ctr": f"{round(float(row.get('ctr', 0) or 0) * 100, 2):g}%",# % format
            "position": round(float(row.get("position", 0) or 0), 2)    # avg position rounded to 2 decimals
        })

    df = pd.DataFrame(rows)

    if not df.empty:
        df[dim_name] = df[dim_name].astype(str).str.lower().str.strip()

    logger.info(f"GSC rows fetched: {len(df)}")
    logger.info(f"Columns returned: {df.columns.tolist()}")
    logger.debug(f"Data length: {len(df)}, \ncolumns: {df.columns.tolist()}, \nshape: {df.shape}")

    return df

def update_daily_page_sheet(service, sheet_name, start_date, end_date, formatted_date, local_file=LOCAL_FILE):
    logger.info(f"Updating sheet: {sheet_name}")

    df = pd.read_excel(local_file, sheet_name=sheet_name, header=[0])
    df_page=fetch_data(service, start_date, end_date, ["page"])
    df_page = filter_trackable_pages(df_page)
    df = build_edit_update_page_df(df,df_page,formatted_date)
    
    with pd.ExcelWriter(
        local_file,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="replace"
    ) as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

def validate_file(client, file_id):
    url = f"https://graph.microsoft.com/v1.0/sites/{SITE_ID}/drives/{DRIVE_ID}/items/{file_id}"
    response = requests.get(url, headers=client.get_headers())

    logger.info("Validate file status: %s", response.status_code)
    logger.info("Validate file body: %s", response.text[:1000])

    return response

    
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def format_daily_sheet_headers(sheet_name, local_file=LOCAL_FILE):
    logger.info(f"Formatting headers for sheet: {sheet_name}")

    wb = load_workbook(local_file)
    ws = wb[sheet_name]

    if ws.max_row < 1:
        logger.info(f"Sheet '{sheet_name}' is empty. Skipping.")
        wb.close()
        return

    # Read existing first row headers
    original_headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]

    # If already formatted, skip
    if ws.max_row >= 2:
        second_row_values = [ws.cell(row=2, column=col).value for col in range(1, ws.max_column + 1)]
        if any(
            str(val).strip().lower() in ["clicks", "impressions", "ctr", "avg position"]
            for val in second_row_values if val
        ):
            logger.info(f"Sheet '{sheet_name}' already formatted. Skipping.")
            wb.close()
            return

    # Store existing data rows (from row 2 onwards)
    data_rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        data_rows.append(list(row))

    # Clear sheet
    ws.delete_rows(1, ws.max_row)

    # Styling
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    subheader_fill = PatternFill("solid", fgColor="EAD1DC")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    col_idx = 1
    i = 0

    while i < len(original_headers):
        header = original_headers[i]
        header_str = str(header).strip() if header is not None else ""

        grouped = False

        # -----------------------------------------
        # Check 4-column grouped format:
        # 11 April_Clicks, 11 April_Impressions, 11 April_CTR, 11 April_Avg Position
        # -----------------------------------------
        if isinstance(header, str) and "_" in header and i + 3 < len(original_headers):
            h1 = str(original_headers[i]).strip()
            h2 = str(original_headers[i + 1]).strip()
            h3 = str(original_headers[i + 2]).strip()
            h4 = str(original_headers[i + 3]).strip()

            if "_" in h1 and "_" in h2 and "_" in h3 and "_" in h4:
                base1, sub1 = h1.rsplit("_", 1)
                base2, sub2 = h2.rsplit("_", 1)
                base3, sub3 = h3.rsplit("_", 1)

                # Special handling for Avg Position (contains underscore + space)
                if "_Avg Position" in h4:
                    base4 = h4.replace("_Avg Position", "")
                    sub4 = "Avg Position"
                else:
                    base4, sub4 = h4.rsplit("_", 1)

                if (
                    base1 == base2 == base3 == base4
                    and sub1.strip().lower() == "clicks"
                    and sub2.strip().lower() == "impressions"
                    and sub3.strip().lower() == "ctr"
                    and sub4.strip().lower() == "avg position"
                ):
                    grouped = True

                    # Parent merged header
                    ws.cell(row=1, column=col_idx, value=base1)
                    ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 3)

                    # Subheaders
                    subheaders = ["Clicks", "Impressions", "CTR", "Avg Position"]
                    for offset, subheader in enumerate(subheaders):
                        ws.cell(row=2, column=col_idx + offset, value=subheader)

                    # Style parent row
                    for c in range(col_idx, col_idx + 4):
                        cell = ws.cell(row=1, column=c)
                        cell.font = bold_font
                        cell.alignment = center_align
                        cell.fill = header_fill
                        cell.border = thin_border

                    # Style subheader row
                    for c in range(col_idx, col_idx + 4):
                        cell = ws.cell(row=2, column=c)
                        cell.font = bold_font
                        cell.alignment = center_align
                        cell.fill = subheader_fill
                        cell.border = thin_border

                    col_idx += 4
                    i += 4

        if grouped:
            continue

        # -----------------------------------------
        # Fallback standalone column (urls, keywords, etc.)
        # -----------------------------------------
        ws.cell(row=1, column=col_idx, value=header_str)
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)

        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border
        ws.cell(row=2, column=col_idx).border = thin_border

        col_idx += 1
        i += 1

    # Write data back starting from row 3
    for row_num, row_data in enumerate(data_rows, start=3):
        for col_num, value in enumerate(row_data, start=1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Freeze top 2 rows
    ws.freeze_panes = "A3"

    # Auto width
    for c in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(c)

        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                max_length = max(max_length, len(str(val)))

        ws.column_dimensions[col_letter].width = max_length + 3

    wb.save(local_file)
    wb.close()

    logger.info(f"Headers formatted successfully for sheet: {sheet_name}")

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def flatten_daily_sheet_headers(sheet_name, local_file=LOCAL_FILE):
    logger.info(f"Flattening headers for sheet: {sheet_name}")
    logger.info(f"Using local file: {local_file}")

    wb = load_workbook(local_file)

    try:
        # Safe sheet lookup (case-insensitive + strip)
        actual_sheet_name = None
        target_norm = sheet_name.strip().lower()

        for s in wb.sheetnames:
            if s.strip().lower() == target_norm:
                actual_sheet_name = s
                break

        if not actual_sheet_name:
            logger.warning(
                f"Sheet '{sheet_name}' not found in local file. "
                f"Available sheets: {wb.sheetnames}. Skipping flatten."
            )
            wb.close()
            return False

        ws = wb[actual_sheet_name]

        # Safety: need at least 2 rows
        if ws.max_row < 2:
            logger.info(f"Sheet '{actual_sheet_name}' has less than 2 rows. Skipping flatten.")
            wb.close()
            return False

        row1 = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
        row2 = [ws.cell(row=2, column=col).value for col in range(1, ws.max_column + 1)]

        # Detect if it's actually grouped for Search Console metrics
        expected_children = {"clicks", "impressions", "ctr", "avg position", "average position", "position"}

        row2_values = {
            str(val).strip().lower()
            for val in row2
            if val is not None and str(val).strip() != ""
        }

        if not any(val in expected_children for val in row2_values):
            logger.info(f"Sheet '{actual_sheet_name}' is already flat. Skipping flatten.")
            wb.close()
            return False

        # Handle merged parent headers by propagating previous non-empty parent
        flat_headers = []
        current_parent = None

        for col_idx in range(1, ws.max_column + 1):
            parent = ws.cell(row=1, column=col_idx).value
            child = ws.cell(row=2, column=col_idx).value

            if parent is not None and str(parent).strip() != "":
                current_parent = str(parent).strip()

            # For vertically merged headers like URLs
            if child is None or str(child).strip() == "":
                flat_headers.append(str(current_parent) if current_parent is not None else "")
            else:
                child = str(child).strip()
                flat_headers.append(f"{current_parent}_{child}" if current_parent else child)

        # Store data rows starting from row 3
        data_rows = []
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            data_rows.append(list(row))

        # Clear sheet
        ws.delete_rows(1, ws.max_row)

        # Write flat headers in row 1
        for col_idx, header in enumerate(flat_headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        # Write data starting from row 2
        for row_num, row_data in enumerate(data_rows, start=2):
            for col_num, value in enumerate(row_data, start=1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Freeze first row
        ws.freeze_panes = "A2"

        # Optional: auto width
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)

            for row_num in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = max_length + 3

        wb.save(local_file)
        logger.info(f"Sheet '{actual_sheet_name}' flattened successfully.")
        wb.close()
        return True

    except Exception as e:
        logger.exception(f"Error while flattening sheet '{sheet_name}': {e}")
        wb.close()
        return False    

def load_all_page_data(start_date, end_date, formatted_date, local_file=LOCAL_FILE, file_id=FILE_ID, sheets_to_update=[]):
    try:
        client = GraphAPIClient()

        # Step 0: Delete existing file
        if os.path.exists(local_file):
            os.remove(local_file)
            logger.info("Old local file removed")

        # Step 1: Download full file locally (temp.xlsx)
        # validate_file(client, FILE_ID)
        download_file(client, local_file, file_id)

        # Step 2: GSC
        service = get_gsc_service()

        for sheet in sheets_to_update:
            # logger.info(f"Flattened headers for sheet: {sheet} and saved to {LOCAL_FILE}")
            flatten_daily_sheet_headers(sheet,local_file)
           
            if sheet == "daily sheet" or sheet == "monthly sheet" or sheet == "weekly sheet":
                update_daily_page_sheet(service, sheet,start_date, end_date,formatted_date,local_file)
    
            # ✅ format this sheet inside temp.xlsx
                format_daily_sheet_headers(sheet,local_file)

            logger.info(f"Formatted headers for sheet: {sheet}")
        
         # =========================
        # 🚀 Step 3: Upload with Retry
        # =========================
        MAX_RETRIES = 3
        upload_success = False
 
        for attempt in range(1, MAX_RETRIES + 1):
            try:
                logger.info(f"Upload attempt {attempt}...")
                upload_file(client, local_file, file_id)
                upload_success = True
                logger.info("Upload successful ✅")
                break
 
            except Exception as e:
                logger.warning(f"Upload attempt {attempt} failed: {str(e)}")
 
                if "locked" in str(e).lower():
                    logger.warning("File is locked on SharePoint. Retrying in 5 seconds...")
                    time.sleep(5)
                else:
                    # Non-retryable error → break immediately
                    break
 
        # =========================
        # 🧹 Step 4: Delete ONLY if upload succeeded
        # =========================
        if upload_success:
            if os.path.exists(local_file):
                os.remove(local_file)
                logger.info("Local temp file deleted 🧹")
        else:
            logger.error("Upload failed after retries. Keeping local file for debugging ⚠️")
 
        logger.info("Pipeline completed successfully ✅" if upload_success else "Pipeline completed with errors ⚠️")
 
    except Exception as e:
        logger.exception(f"Pipeline failed: {str(e)}")

def main():
    logger.info("Starting SEO pipeline...")
    target_date = date.today() - timedelta(days=2) 
    start_date = target_date
    end_date = target_date
    formatted_date = target_date.strftime("%d %B")
    
    sheets_to_update = [
           "daily sheet",
        ]
    
    load_all_page_data(start_date, end_date, formatted_date, LOCAL_FILE, FILE_ID, sheets_to_update)
if __name__ == "__main__":
    main()