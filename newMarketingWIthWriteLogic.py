import math
import os
import time
import io
import requests
import pandas as pd
import logging
from datetime import date, timedelta
from dotenv import load_dotenv
from google.oauth2 import service_account
from googleapiclient.discovery import build
from openpyxl import load_workbook
import re

# =========================
# 📝 LOGGING CONFIG
# =========================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("app.log", mode="a", encoding="utf-8")
    ]
)

logger = logging.getLogger(__name__)

# =========================
# 🔐 ENV
# =========================
load_dotenv()

TENANT_ID = os.getenv("TENANT_ID")
CLIENT_ID = os.getenv("CLIENT_ID")
CLIENT_SECRET = os.getenv("CLIENT_SECRET")

SITE_ID = os.getenv("SITE_ID")
DRIVE_ID = os.getenv("DRIVE_ID")
FILE_ID = os.getenv("COMBINED_FILE_ID")

SITE_URL = os.getenv("GSC_SITE_URL")
KEY_FILE = os.getenv("GSC_KEY_FILE")

LOCAL_FILE = "temp.xlsx"

COUNTRY_MAP = {
    "South Africa": "zaf",
    "Brazil": "bra",
    "Turkey": "tur",
    "Nigeria": "nga",
    "Kenya": "ken",
    "India": "ind",
}

# =========================
# 🔑 GRAPH CLIENT
# =========================
class GraphAPIClient:
    def __init__(self):
        self.access_token = None
        self.expiry_time = 0

    def get_token(self):
        if self.access_token and time.time() < self.expiry_time:
            logger.debug("Using cached token")
            return self.access_token

        logger.info("Fetching new Graph API token...")
        token_url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/token"

        data = {
            "grant_type": "client_credentials",
            "client_id": CLIENT_ID,
            "client_secret": CLIENT_SECRET,
            "resource": "https://graph.microsoft.com/"
        }

        response = requests.post(token_url, data=data).json()

        if "access_token" not in response:
            raise Exception(response)

        self.access_token = response["access_token"]
        self.expiry_time = time.time() + int(response["expires_in"]) - 60

        logger.info("Token fetched successfully")
        return self.access_token

    def get_headers(self):
        return {"Authorization": f"Bearer {self.get_token()}"}

# =========================
# 📂 DOWNLOAD FULL FILE
# =========================
# def download_file(client, local_file=LOCAL_FILE,file_id=FILE_ID):
#     logger.info("Downloading full Excel file from SharePoint...")

#     url = f"https://graph.microsoft.com/v1.0/sites/{SITE_ID}/drives/{DRIVE_ID}/items/{file_id}/content"
#     response = requests.get(url, headers=client.get_headers())

#     if response.status_code != 200:
#         raise Exception(response.text)

#     with open(local_file, "wb") as f:
#         f.write(response.content)

#     logger.info("File downloaded successfully")

def download_file(client, local_file=LOCAL_FILE, file_id=FILE_ID):
    logger.info("Downloading full Excel file from SharePoint...")
    logger.info("Using FILE_ID: %s", file_id)
    logger.info("Using LOCAL_FILE: %s", local_file)

    url = f"https://graph.microsoft.com/v1.0/sites/{SITE_ID}/drives/{DRIVE_ID}/items/{file_id}/content"
    
    
    response = requests.get(url, headers=client.get_headers(), allow_redirects=False)

    # logger.info("Download status code: %s", response.status_code)
    # logger.info("Download headers: %s", dict(response.headers))
    # logger.info("Download response text: %s", response.text[:500])

    if response.status_code == 200:
        with open(local_file, "wb") as f:
            f.write(response.content)
        logger.info("File downloaded successfully")
        return

    # Handle Graph redirect to actual file URL
    if response.status_code in (301, 302, 303, 307, 308):
        download_url = response.headers.get("Location")
        if not download_url:
            raise Exception(f"Redirect received but no Location header. Status={response.status_code}")

        logger.info("Following redirect to actual download URL...")

        # IMPORTANT: redirected URL usually should be called WITHOUT auth header
        file_response = requests.get(download_url, stream=True)

        logger.info("Redirected download status code: %s", file_response.status_code)

        if file_response.status_code != 200:
            raise Exception(
                f"Redirected download failed. "
                f"Status={file_response.status_code}, Body={file_response.text[:500]}"
            )

        with open(local_file, "wb") as f:
            for chunk in file_response.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)

        logger.info("File downloaded successfully after redirect")
        return

    raise Exception(
        f"Download failed. Status={response.status_code}, Body={response.text[:500]}"
    )

# =========================
# 📤 UPLOAD FILE
# =========================
def upload_file(client, local_file=LOCAL_FILE,file_id=FILE_ID):
    logger.info("Uploading updated file to SharePoint...")

    url = f"https://graph.microsoft.com/v1.0/sites/{SITE_ID}/drives/{DRIVE_ID}/items/{file_id}/content"

    with open(local_file, "rb") as f:
        data = f.read()

    headers = client.get_headers()
    headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    response = requests.put(url, headers=headers, data=data)

    if response.status_code not in [200, 201]:
        raise Exception(response.text)

    logger.info("Upload successful ✅")

# =========================
# 🔧 FIX MULTIINDEX
# =========================
def fix_multiindex_columns(df):
    new_cols = []
    last_main = None

    for main, sub in df.columns:
        main = str(main).strip()
        sub = str(sub).strip().lower()

        if "unnamed" in main.lower():
            main = last_main
        else:
            last_main = main

        if "unnamed" in sub:
            new_cols.append(main)
        else:
            new_cols.append(f"{main}_{sub}")

    df.columns = new_cols
    return df

# =========================
# 🔎 COLUMN FINDER
# =========================
def get_column(df, keyword):
    for col in df.columns:
        if keyword in str(col).lower():
            return col
    return None

# =========================
# 🔎 GSC
# =========================
def get_gsc_service():
    logger.info("Initializing GSC service...")
    credentials = service_account.Credentials.from_service_account_file(
        KEY_FILE,
        scopes=['https://www.googleapis.com/auth/webmasters.readonly']
    )
    return build('searchconsole', 'v1', credentials=credentials)

def fetch_page_data_by_country(service, site_url, start_date, end_date, country_code, row_limit=25000):
    """
    Fetch page-wise clicks and average position for a specific country from GSC.

    Args:
        service: Authorized Google Search Console service object
        site_url: GSC property URL (must exactly match verified property)
        start_date: datetime.date or string YYYY-MM-DD
        end_date: datetime.date or string YYYY-MM-DD
        country_code: ISO country code (e.g. 'ind', 'usa', 'gbr')
        row_limit: max rows per request (max 25000 per request)

    Returns:
        pandas.DataFrame with columns: page, clicks, position
    """

    logger.info(f"Fetching GSC page data from {start_date} to {end_date} for country={country_code}")

    all_rows = []
    start_row = 0

    while True:
        body = {
            "startDate": str(start_date),
            "endDate": str(end_date),
            "dimensions": ["page"],
            "rowLimit": row_limit,
            "startRow": start_row,
            "type": "web",
            "dimensionFilterGroups": [
                {
                    "filters": [
                        {
                            "dimension": "country",
                            "operator": "equals",
                            "expression": country_code.lower()
                        }
                    ]
                }
            ]
        }

        response = service.searchanalytics().query(
            siteUrl=site_url,
            body=body
        ).execute()

        rows = response.get("rows", [])
        if not rows:
            break

        for row in rows:
            all_rows.append({
                "page": row["keys"][0],
                "clicks": row.get("clicks", 0),
                "position": math.ceil(float(row.get("position") or 0))
            })

        logger.info(f"Fetched {len(rows)} page rows at startRow={start_row}")

        # If fewer than row_limit returned, no more pages
        if len(rows) < row_limit:
            break

        start_row += row_limit

    df = pd.DataFrame(all_rows)

    if not df.empty:
        df["page"] = (
            df["page"]
            .astype(str)
            .str.lower()
            .str.strip()
            .str.replace("\u00a0", "", regex=False)
        )

    logger.info(f"Total GSC page rows fetched: {len(df)}")
    logger.info(f"Columns returned: {df.columns.tolist()}")

    return df

def fetch_keyword_data_by_country(service, site_url, start_date, end_date, country_code, row_limit=25000):
    """
    Fetch keyword-wise clicks and average position for a specific country from GSC.

    Args:
        service: Authorized Google Search Console service object
        site_url: GSC property URL (must exactly match verified property)
        start_date: datetime.date or string YYYY-MM-DD
        end_date: datetime.date or string YYYY-MM-DD
        country_code: ISO country code (e.g. 'ind', 'usa', 'gbr')
        row_limit: max rows per request (max 25000 per request)

    Returns:
        pandas.DataFrame with columns: query, clicks, impressions, ctr, position
    """

    logger.info(f"Fetching GSC keyword data from {start_date} to {end_date} for country={country_code}")

    all_rows = []
    start_row = 0

    while True:
        body = {
            "startDate": str(start_date),
            "endDate": str(end_date),
            "dimensions": ["query"],
            "rowLimit": row_limit,
            "startRow": start_row,
            "type": "web",
            "dimensionFilterGroups": [
                {
                    "filters": [
                        {
                            "dimension": "country",
                            "operator": "equals",
                            "expression": country_code.lower()
                        }
                    ]
                }
            ]
        }
 
        response = service.searchanalytics().query(
            siteUrl=site_url,
            body=body
        ).execute()

        rows = response.get("rows", [])
        if not rows:
            break

        for row in rows:
            all_rows.append({
                "query": row["keys"][0],
                "clicks": row.get("clicks", 0),
                # "impressions": row.get("impressions", 0),
                # "ctr": row.get("ctr", 0),
                "position": math.ceil(float(row.get("position") or 0))           
            })

        logger.info(f"Fetched {len(rows)} rows at startRow={start_row}")

        # If fewer than row_limit returned, no more pages
        if len(rows) < row_limit:
            break

        start_row += row_limit

    df = pd.DataFrame(all_rows)

    if not df.empty:
        df["query"] = df["query"].astype(str).str.lower().str.strip()

    logger.info(f"Total GSC keyword rows fetched: {len(df)}")
    logger.info(f"Columns returned: {df.columns.tolist()}")

    return df

def fetch_data(service, start_date, end_date, dimension):
    logger.info(f"Fetching GSC data from {start_date} to {end_date}")

    response = service.searchanalytics().query(
        siteUrl=SITE_URL,
        body={
            "startDate": str(start_date),
            "endDate": str(end_date),
            "dimensions": dimension,
            "rowLimit": 25000
        }
    ).execute()

    dim_name = dimension[0]  # "query" OR "page"

    rows = []
    for row in response.get("rows", []):
        rows.append({
            dim_name: row["keys"][0],
            "clicks": row["clicks"],
            "position": math.ceil(float(row["position"] or 0))
        })

    df = pd.DataFrame(rows)

    if not df.empty:
        df[dim_name] = df[dim_name].astype(str).str.lower().str.strip()

    logger.info(f"GSC rows fetched: {len(df)}")
    logger.info(f"Columns returned: {df.columns.tolist()}")
    logger.debug(f"Data length: {len(df)}, \ncolumns: {df.columns.tolist()}, \nshape: {df.shape}")
    return df

# =========================
# 🔄 HELPER METHOD FOR PAGES
# =========================

def build_page_df(df_page_sheet, df_gsc_page, end_date):
    # df_page_sheet = df_page_sheet.copy()
    # df_gsc_page = df_gsc_page.copy()
 
    # ==========================================
    # 1) Flatten old grouped-header columns
    # Example:
    # urls | 2026-03-20 00:00:00 | Unnamed: 2
    # =>
    # urls | 3/20/2026_Rank | 3/20/2026_Traffic
    # ==========================================
    cols = list(df_page_sheet.columns)
    new_cols = []
    i = 0
 
    while i < len(cols):
        col = str(cols[i]).strip()
 
        # Keep URL column as-is
        if col.lower() == "urls":
            new_cols.append("urls")
            i += 1
            continue
 
        # Old grouped-header style:
        # first col = date, second col = Unnamed => treat as Rank + Traffic
        if "Unnamed:" not in col:
            next_col = str(cols[i + 1]).strip() if i + 1 < len(cols) else None
 
            if next_col and "Unnamed:" in next_col:
                # Convert date-like header to m/d/YYYY format
                try:
                    dt = pd.to_datetime(col)
                    date_label = f"{dt.month}/{dt.day}/{dt.year}"
                except Exception:
                    # fallback if parsing fails
                    date_label = col.split(" ")[0]
 
                new_cols.append(f"{date_label}_Rank")
                new_cols.append(f"{date_label}_Traffic")
                i += 2
                continue
 
        # Already flat column -> keep as-is
        new_cols.append(col)
        i += 1
 
    # Apply flattened column names
    if len(new_cols) == len(df_page_sheet.columns):
        df_page_sheet.columns = new_cols
 
    # ==========================================
    # 2) Create normalized temp copy for URL matching only
    # ==========================================
    df_temp = df_page_sheet.copy()
    df_temp.columns = (
        df_temp.columns.astype(str)
        .str.strip()
        .str.lower()
        .str.replace(" ", "_")
    )

    logger.info("tdf_columns : %s", df_temp.columns)
 
    # Find URL column safely
    url_col = "urls"
    if url_col not in df_temp.columns:
        raise KeyError(f"Expected column '{url_col}' in Page sheet. Found: {list(df_temp.columns)}")
 
    # Normalize URLs in page sheet
    df_temp[url_col] = df_temp[url_col].astype(str).str.lower().str.strip()
 
    # Normalize GSC page URLs
    df_gsc_page["page"] = df_gsc_page["page"].astype(str).str.lower().str.strip()
 
    # ==========================================
    # 3) Build page-level metrics from GSC
    # ==========================================
    page_metrics = (
        df_gsc_page.groupby("page", as_index=False)
        .agg(
            traffic=("clicks", "sum"),
            rank=("position", "min")
        )
    )
 
    metrics_lookup = page_metrics.set_index("page")
 
    # ==========================================
    # 4) New column names in SAME flat format
    # Example:
    # 3/30/2026_Rank
    # 3/30/2026_Traffic
    # ==========================================
    # date_label = f"{end_date.month}/{end_date.day}/{end_date.year}"
    rank_col = f"{end_date}_Rank"
    traffic_col = f"{end_date}_Traffic"
 
    # Optional: overwrite if already exists (safe rerun)
    df_page_sheet[rank_col] = df_temp[url_col].map(metrics_lookup["rank"])
    df_page_sheet[traffic_col] = df_temp[url_col].map(metrics_lookup["traffic"])
 
    return df_page_sheet 

# =========================
# 🔄 UPDATE ONE SHEET
# =========================
def update_kw_sheet(service,sheet_name, start_date, end_date, formatted_date):
    logger.info(f"Updating sheet: {sheet_name}")
    
    df_query = fetch_data(service, start_date, end_date, ["query"])

    df = pd.read_excel(LOCAL_FILE, sheet_name=sheet_name, header=[0])
    # df = fix_multiindex_columns(df)

    primary_col = get_column(df, "primary")
    secondary_col = get_column(df, "secondary")

    rank_col = f"{formatted_date}_rank"
    traffic_col = f"{formatted_date}_traffic"

    if rank_col not in df.columns:
        df[rank_col] = None
    if traffic_col not in df.columns:
        df[traffic_col] = None

    for idx, row in df.iterrows():
        primary = row.get(primary_col, "")
        secondary = row.get(secondary_col, "")

        primary = "" if pd.isna(primary) else str(primary).lower()
        secondary = "" if pd.isna(secondary) else str(secondary).lower()

        keywords = [k.strip() for k in (primary + "," + secondary).split(",") if k.strip()]

        if not keywords:
            continue

        pattern = "|".join(map(re.escape, keywords))

        matched = df_query[df_query["query"].str.contains(pattern, na=False)]

        if not matched.empty:
            best = matched.sort_values("clicks", ascending=False).iloc[0]
            df.at[idx, rank_col] = best["position"]
            df.at[idx, traffic_col] = best["clicks"]

    # ✅ SAFE WRITE (modern pandas way)
    with pd.ExcelWriter(
        LOCAL_FILE,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="replace"
    ) as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    logger.info(f"Sheet '{sheet_name}' updated successfully")

def update_page_sheet(service, sheet_name, start_date, end_date, formatted_date):
    logger.info(f"Updating sheet: {sheet_name}")

    df = pd.read_excel(LOCAL_FILE, sheet_name=sheet_name, header=[0])
    df_page=fetch_data(service, start_date, end_date, ["page"])
    df = build_page_df(df,df_page,formatted_date)
    
    with pd.ExcelWriter(
        LOCAL_FILE,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="replace"
    ) as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    logger.info(f"Sheet '{sheet_name}' updated successfully")
    
def update_country_sheet(service, sheet_name, start_date, end_date, formatted_date):
    logger.info(f"Updating country sheet: {sheet_name}")

    country_code = COUNTRY_MAP.get(sheet_name)
    if not country_code:
        logger.warning(f"No country code mapping found for sheet '{sheet_name}'. Skipping.")
        return

    # 1) Read existing sheet
    df = pd.read_excel(LOCAL_FILE, sheet_name=sheet_name, header=0)

    keyword_col = "keywords"
    if keyword_col not in df.columns:
        logger.warning(
            f"'{keyword_col}' column not found in sheet '{sheet_name}'. "
            f"Found columns: {df.columns.tolist()}. Skipping."
        )
        return

    # Normalize sheet keywords
    df[keyword_col] = df[keyword_col].astype(str).str.lower().str.strip()

    # 2) Fetch GSC data for this country
    # service = get_gsc_service()
    df_country = fetch_keyword_data_by_country(
        service,
        SITE_URL,
        start_date,
        end_date,
        country_code
    )

    # Dynamic output columns
    rank_col = f"{formatted_date}_rank"
    traffic_col = f"{formatted_date}_traffic"

    # Remove old columns for same date if already present (safe re-run)
    df = df.drop(columns=[rank_col, traffic_col], errors="ignore")

    # 3) If no GSC data, still add empty columns and write back
    if df_country.empty:
        logger.warning(f"No GSC data found for country '{sheet_name}' ({country_code})")

        df[rank_col] = None
        df[traffic_col] = None

        with pd.ExcelWriter(
            LOCAL_FILE,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace"
        ) as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        logger.info(f"Country sheet '{sheet_name}' updated with empty rank/traffic columns")
        return

    # Normalize GSC query
    df_country["query"] = df_country["query"].astype(str).str.lower().str.strip()

    # 4) Prepare GSC mapping dataframe
    df_country_map = df_country[["query", "position", "clicks"]].copy()
    df_country_map = df_country_map.rename(columns={
        "query": "keywords",
        "position": rank_col,
        "clicks": traffic_col
    })

    # If same keyword appears multiple times, keep the one with highest clicks
    df_country_map = (
        df_country_map.sort_values(traffic_col, ascending=False)
        .drop_duplicates(subset=["keywords"], keep="first")
    )

    # 5) Merge existing sheet with GSC data
    df_updated = df.merge(df_country_map, on="keywords", how="left")

    # 6) Write updated sheet back
    with pd.ExcelWriter(
        LOCAL_FILE,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="replace"
    ) as writer:
        df_updated.to_excel(writer, sheet_name=sheet_name, index=False)

    logger.info(f"Country sheet '{sheet_name}' updated successfully")
    
def build_country_page_df(df_page_sheet, df_gsc_page, formatted_date):
    df_page_sheet = df_page_sheet.copy()
    df_gsc_page = df_gsc_page.copy()

    # Normalize page sheet URLs
    df_page_sheet["urls"] = (
        df_page_sheet["urls"]
        .astype(str)
        .str.strip()
        .str.lower()
        .str.replace("\u00a0", "", regex=False)  # remove non-breaking spaces
    )

    # Dynamic columns
    rank_col = f"{formatted_date}_Rank"
    traffic_col = f"{formatted_date}_Traffic"

    # Safe rerun: remove if already exists
    df_page_sheet = df_page_sheet.drop(columns=[rank_col, traffic_col], errors="ignore")

    # If no GSC data, still add empty cols
    if df_gsc_page.empty:
        logger.warning("No page-level GSC data found for this country")
        df_page_sheet[rank_col] = None
        df_page_sheet[traffic_col] = None
        return df_page_sheet

    # Normalize GSC page URLs
    df_gsc_page["page"] = (
        df_gsc_page["page"]
        .astype(str)
        .str.strip()
        .str.lower()
        .str.replace("\u00a0", "", regex=False)
    )

    # Aggregate page-level metrics
    page_metrics = (
        df_gsc_page.groupby("page", as_index=False)
        .agg(
            traffic=("clicks", "sum"),
            rank=("position", "min")   # same as your current logic
        )
    )

    # Rename to final output cols
    page_metrics = page_metrics.rename(columns={
        "page": "urls",
        "rank": rank_col,
        "traffic": traffic_col
    })

    # Merge on urls
    df_updated = df_page_sheet.merge(
        page_metrics[["urls", rank_col, traffic_col]],
        on="urls",
        how="left"
    )

    return df_updated    
    
def update_country_page_sheet(service, sheet_name, start_date, end_date, formatted_date):
    logger.info(f"Updating country page sheet: {sheet_name}")

    # Example: India_page -> India
    base_sheet_name = sheet_name.replace(" Page", "").strip()
    
    logger.info(f"Base sheet name extracted: '{base_sheet_name}' from '{sheet_name}'")

    country_code = COUNTRY_MAP.get(base_sheet_name)
    if not country_code:
        logger.warning(
            f"No country code mapping found for sheet '{sheet_name}' "
            f"(base key: '{base_sheet_name}'). Skipping."
        )
        return

    # Read existing page sheet
    df_page_sheet = pd.read_excel(LOCAL_FILE, sheet_name=sheet_name, header=0)

    # Validate urls column
    if "urls" not in df_page_sheet.columns:
        logger.warning(
            f"'urls' column not found in sheet '{sheet_name}'. "
            f"Found columns: {df_page_sheet.columns.tolist()}. Skipping."
        )
        return

    # Fetch GSC page data filtered by country
    # service = get_gsc_service()
    df_gsc_page = fetch_page_data_by_country(
        service=service,
        site_url=SITE_URL,
        start_date=start_date,
        end_date=end_date,
        country_code=country_code
    )

    # Build updated page sheet
    df_updated = build_country_page_df(df_page_sheet, df_gsc_page, formatted_date)

    # Write back
    with pd.ExcelWriter(
        LOCAL_FILE,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="replace"
    ) as writer:
        df_updated.to_excel(writer, sheet_name=sheet_name, index=False)

    logger.info(f"Country page sheet '{sheet_name}' updated successfully")    
    
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def format_sheet_headers(sheet_name,local_file=LOCAL_FILE):
    """
    Converts flat headers like:
        urls | 05 April_Rank | 05 April_Traffic | keywords
    into:
        Row 1: urls | 05 April (merged) | keywords
        Row 2:      | Rank | Traffic    |

    Generic:
    - Any valid grouped pair like DATE_Rank + DATE_Traffic becomes grouped
    - Everything else remains a normal vertically merged header
    """
    logger.info(f"Formatting headers for sheet: {sheet_name}")

    wb = load_workbook(local_file)
    ws = wb[sheet_name]

    if ws.max_row < 1:
        logger.info(f"Sheet '{sheet_name}' is empty. Skipping.")
        wb.close()
        return

    # Read existing first row headers
    original_headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]

    # If already formatted, skip
    if ws.max_row >= 2:
        second_row_values = [ws.cell(row=2, column=col).value for col in range(1, ws.max_column + 1)]
        if any(str(val).strip().lower() in ["rank", "traffic"] for val in second_row_values if val):
            logger.info(f"Sheet '{sheet_name}' already formatted. Skipping.")
            wb.close()
            return

    # Store existing data rows (from row 2 onwards)
    data_rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        data_rows.append(list(row))

    # Clear sheet
    ws.delete_rows(1, ws.max_row)

    # Styling
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    subheader_fill = PatternFill("solid", fgColor="EAD1DC")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    col_idx = 1
    i = 0

    while i < len(original_headers):
        header = original_headers[i]

        # Normalize header
        header_str = str(header).strip() if header is not None else ""

        # -----------------------------
        # Check if this is a grouped pair
        # Example: 05 April_Rank + 05 April_Traffic
        # -----------------------------
        grouped = False

        if isinstance(header, str) and "_" in header and i + 1 < len(original_headers):
            base_name, sub_name = header.rsplit("_", 1)
            next_header = original_headers[i + 1]

            if isinstance(next_header, str) and "_" in next_header:
                next_base, next_sub = next_header.rsplit("_", 1)

                # Valid pair only if same base and expected subheaders
                if (
                    base_name == next_base
                    and sub_name.strip().lower() == "rank"
                    and next_sub.strip().lower() == "traffic"
                ):
                    grouped = True

                    # Parent header
                    ws.cell(row=1, column=col_idx, value=base_name)
                    ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 1)

                    # Subheaders
                    ws.cell(row=2, column=col_idx, value="Rank")
                    ws.cell(row=2, column=col_idx + 1, value="Traffic")

                    # Style parent
                    parent_cell = ws.cell(row=1, column=col_idx)
                    parent_cell.font = bold_font
                    parent_cell.alignment = center_align
                    parent_cell.fill = header_fill
                    parent_cell.border = thin_border
                    ws.cell(row=1, column=col_idx + 1).border = thin_border

                    # Style subheaders
                    for c in [col_idx, col_idx + 1]:
                        cell = ws.cell(row=2, column=c)
                        cell.font = bold_font
                        cell.alignment = center_align
                        cell.fill = subheader_fill
                        cell.border = thin_border

                    col_idx += 2
                    i += 2

        if grouped:
            continue

        # -----------------------------
        # Fallback = normal standalone column
        # Handles urls, keywords, blog urls, query, page, etc.
        # -----------------------------
        ws.cell(row=1, column=col_idx, value=header_str)
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)

        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border
        ws.cell(row=2, column=col_idx).border = thin_border

        col_idx += 1
        i += 1

    # Write data back starting from row 3
    for row_num, row_data in enumerate(data_rows, start=3):
        for col_num, value in enumerate(row_data, start=1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Freeze top 2 rows
    ws.freeze_panes = "A3"

    # Optional: auto width
    for c in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(c)

        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                max_length = max(max_length, len(str(val)))

        ws.column_dimensions[col_letter].width = max_length + 3

    wb.save(local_file)
    wb.close()

    logger.info(f"Headers formatted successfully for sheet: {sheet_name}")    
    
#Reverse of header formatting

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def flatten_sheet_headers(sheet_name,local_file=LOCAL_FILE):
    """
    Converts 2-row grouped headers like:
        Row 1: urls | 05 April | 05 April
        Row 2:      | Rank     | Traffic

    into flat single-row headers:
        urls | 05 April_Rank | 05 April_Traffic

    Works directly on LOCAL_FILE (temp.xlsx)
    """
    logger.info(f"Flattening headers for sheet: {sheet_name}")
    logger.info(f"Using local file: {local_file}")
    wb = load_workbook(local_file)
    ws = wb[sheet_name]

    # Safety: need at least 2 rows to flatten
    if ws.max_row < 2:
        logger.info(f"Sheet '{sheet_name}' has less than 2 rows. Skipping flatten.")
        wb.close()
        return

    row1 = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    row2 = [ws.cell(row=2, column=col).value for col in range(1, ws.max_column + 1)]

    # Detect if it's actually grouped
    if not any(val in ["Rank", "Traffic"] for val in row2 if val):
        logger.info(f"Sheet '{sheet_name}' is already flat. Skipping flatten.")
        wb.close()
        return

    # Handle merged parent headers by propagating previous non-empty parent
    flat_headers = []
    current_parent = None

    for col_idx in range(1, ws.max_column + 1):
        parent = ws.cell(row=1, column=col_idx).value
        child = ws.cell(row=2, column=col_idx).value

        if parent is not None:
            current_parent = parent

        # urls or any vertically merged single header
        if child is None:
            flat_headers.append(str(current_parent) if current_parent is not None else "")
        else:
            flat_headers.append(f"{current_parent}_{child}")

    # Store data rows starting from row 3
    data_rows = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        data_rows.append(list(row))

    # Clear sheet
    ws.delete_rows(1, ws.max_row)

    # Write flat headers in row 1
    for col_idx, header in enumerate(flat_headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Write data starting from row 2
    for row_num, row_data in enumerate(data_rows, start=2):
        for col_num, value in enumerate(row_data, start=1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Freeze first row
    ws.freeze_panes = "A2"

    # Optional: auto width
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)

        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 3

    wb.save(local_file)
    wb.close()

    logger.info(f"Sheet '{sheet_name}' flattened successfully.")    
# =========================
# 🚀 MAIN
# =========================
def main():
    logger.info("Starting SEO pipeline...")

    try:
        client = GraphAPIClient()

        # Step 0: Delete existing file
        if os.path.exists(LOCAL_FILE):
            os.remove(LOCAL_FILE)
            logger.info("Old local file removed")

        # Step 1: Download full file locally (temp.xlsx)
        download_file(client)

        # Step 2: GSC
        service = get_gsc_service()

        end_date = date.today() - timedelta(days=1)
        start_date = end_date - timedelta(days=6)
        formatted_date = start_date.strftime("%d %B")+" - "+end_date.strftime("%d %B")

        # df_query = fetch_data(service, start_date, end_date, ["query"])
        # df_page = fetch_data(service, start_date, end_date, ["page"])

        # logger.info("pages %s", df_page)

        sheets_to_update = [
            "globalKws",
            "Page sheet",
            # "India",
            "South Africa",
            "South Africa Page",
            "Brazil",
            "Brazil Page",
            "Turkey",
            "Turkey Page",
            "Nigeria",
            "Nigeria Page",
            "Kenya",
            "Kenya Page"
        ]

        for sheet in sheets_to_update:
            
            flatten_sheet_headers(sheet)

            if sheet == 'demoSheet':
                update_kw_sheet(service, sheet,start_date, end_date,formatted_date)

            elif sheet == "Page sheet":
                update_page_sheet(service, sheet,start_date, end_date,formatted_date)

            elif sheet.endswith(" Page"):
                update_country_page_sheet(service, sheet, start_date, end_date, formatted_date)

            else:
                update_country_sheet(service, sheet, start_date, end_date, formatted_date)

            # ✅ format this sheet inside temp.xlsx
            format_sheet_headers(sheet)

            logger.info(f"Formatted headers for sheet: {sheet}")
        
         # =========================
        # 🚀 Step 3: Upload with Retry
        # =========================
        MAX_RETRIES = 3
        upload_success = False
 
        for attempt in range(1, MAX_RETRIES + 1):
            try:
                logger.info(f"Upload attempt {attempt}...")
                upload_file(client)
                upload_success = True
                logger.info("Upload successful ✅")
                break
 
            except Exception as e:
                logger.warning(f"Upload attempt {attempt} failed: {str(e)}")
 
                if "locked" in str(e).lower():
                    logger.warning("File is locked on SharePoint. Retrying in 5 seconds...")
                    time.sleep(5)
                else:
                    # Non-retryable error → break immediately
                    break
 
        # =========================
        # 🧹 Step 4: Delete ONLY if upload succeeded
        # =========================
        if upload_success:
            if os.path.exists(LOCAL_FILE):
                os.remove(LOCAL_FILE)
                logger.info("Local temp file deleted 🧹")
        else:
            logger.error("Upload failed after retries. Keeping local file for debugging ⚠️")
 
        logger.info("Pipeline completed successfully ✅" if upload_success else "Pipeline completed with errors ⚠️")
 
    except Exception as e:
        logger.exception(f"Pipeline failed: {str(e)}")
if __name__ == "__main__":
    main()