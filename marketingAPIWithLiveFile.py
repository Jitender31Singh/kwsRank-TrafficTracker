import os
import time
import io
import requests
import pandas as pd
import logging
from datetime import date, timedelta
from dotenv import load_dotenv
from google.oauth2 import service_account
from googleapiclient.discovery import build
from openpyxl.utils import get_column_letter

# =========================
# 📝 LOGGING CONFIG
# =========================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("app.log", mode="a")
    ]
)

logger = logging.getLogger(__name__)

# =========================
# 🔐 LOAD ENV VARIABLES
# =========================
load_dotenv()

TENANT_ID = os.getenv("TENANT_ID")
CLIENT_ID = os.getenv("CLIENT_ID")
CLIENT_SECRET = os.getenv("CLIENT_SECRET")

SITE_ID = os.getenv("SITE_ID")
DRIVE_ID = os.getenv("DRIVE_ID")
FILE_ID = os.getenv("FILE_ID")

SITE_URL = os.getenv("GSC_SITE_URL")
KEY_FILE = os.getenv("GSC_KEY_FILE")

OUTPUT_FILE = os.getenv("OUTPUT_FILE", "output.xlsx")

# =========================
# 🔑 GRAPH API TOKEN MANAGER
# =========================
class GraphAPIClient:
    def __init__(self):
        self.access_token = None
        self.expiry_time = 0

    def get_token(self):
        if self.access_token and time.time() < self.expiry_time:
            logger.debug("Using cached access token")
            return self.access_token

        logger.info("Fetching new access token...")

        token_url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"

        data = {
            "grant_type": "client_credentials",
            "client_id": CLIENT_ID,
            "client_secret": CLIENT_SECRET,
            "scope": "https://graph.microsoft.com/.default"
        }

        response = requests.post(token_url, data=data).json()

        if "access_token" not in response:
            logger.error(f"Token fetch failed: {response}")
            raise Exception(f"Token Error: {response}")

        self.access_token = response["access_token"]
        self.expiry_time = time.time() + int(response["expires_in"]) - 60

        logger.info("Access token fetched successfully")

        return self.access_token

    def get_headers(self):
        return {
            "Authorization": f"Bearer {self.get_token()}"
        }
    # def get_headers(self):
    #     return {
    #         "Authorization": f"Bearer {self.access_token}"
    #     }

# =========================
# 📂 FETCH FILE FROM SHAREPOINT
# =========================
def fetch_excel_from_sharepoint(client):
    logger.info("Fetching Excel file from SharePoint...")

    file_url = f"https://graph.microsoft.com/v1.0/sites/{SITE_ID}/drives/{DRIVE_ID}/items/{FILE_ID}/content"

    response = requests.get(file_url, headers=client.get_headers())

    if response.status_code != 200:
        logger.error(f"SharePoint Fetch Error: {response.text}")
        raise Exception(f"SharePoint Fetch Error: {response.text}")

    logger.info("File fetched successfully from SharePoint")

    df = pd.read_excel(io.BytesIO(response.content))
    logger.info(f"Loaded DataFrame with shape: {df.shape}")
    
    print(df.head())  # Debug: Print first few rows of the DataFrame

    return df

# =========================
# 🔎 GOOGLE SEARCH CONSOLE SETUP
# =========================
def get_gsc_service():
    logger.info("Initializing Google Search Console service...")

    SCOPES = ['https://www.googleapis.com/auth/webmasters.readonly']

    credentials = service_account.Credentials.from_service_account_file(
        KEY_FILE, scopes=SCOPES
    )

    logger.info("GSC service initialized successfully")

    return build('searchconsole', 'v1', credentials=credentials)

# =========================
# 🔥 FETCH GSC DATA
# =========================

def fetch_all_page_traffic_rank(service, start_date, end_date):
    logger.info("Fetching GSC traffic and rank for ALL page URLs...")

    # Reuse your existing generic method
    df_page = fetch_data(service, start_date, end_date, ["page"])

    if df_page.empty:
        logger.warning("No page data returned from GSC.")
        return pd.DataFrame(columns=["blog_url", "traffic", "rank"])

    # Normalize page URLs
    df_page["page"] = df_page["page"].astype(str).str.strip().str.lower()

    # Aggregate by page URL
    df_result = (
        df_page.groupby("page", as_index=False)
        .agg(
            traffic=("clicks", "sum"),
            rank=("position", "min")
        )
        .rename(columns={"page": "blog_url"})
    )

    logger.info(f"Fetched traffic/rank for {len(df_result)} page URLs from GSC")

    return df_result
def fetch_data(service, start_date, end_date, dimensions):
    logger.info(f"Fetching GSC data for dimensions: {dimensions}")

    response = service.searchanalytics().query(
        siteUrl=SITE_URL,
        body={
            "startDate": str(start_date),
            "endDate": str(end_date),
            "dimensions": dimensions,
            "rowLimit": 25000
        }
    ).execute()

    rows = []
    for row in response.get("rows", []):
        record = {
            "traffic": row["clicks"],
            # "impressions": row["impressions"],
            # "ctr": row["ctr"],
            "rank": row["position"]
        }

        for i, dim in enumerate(dimensions):
            record[dim] = row["keys"][i]

        rows.append(record)

    df = pd.DataFrame(rows)
    logger.info(f"GSC data fetched: {len(df)} rows for {dimensions}")

    return df

# =========================
# 🚀 MAIN EXECUTION
# =========================
def main():
    logger.info(" Starting SEO pipeline...")

    try:
        graph_client = GraphAPIClient()

        sheets_to_process = ["Sheet1"]

        service = get_gsc_service()
        end_date = date.today() - timedelta(days=1)
        start_date = end_date - timedelta(days=7)

        df_query = fetch_data(service, start_date, end_date, ["query"])
        df_gsc_page = fetch_data(service, start_date, end_date, ["page"])
        df_country = fetch_data(service, start_date, end_date, ["country"])
        df_device = fetch_data(service, start_date, end_date, ["device"])
        df_query_page = fetch_data(service, start_date, end_date, ["query", "page"])

        # Build final page dataframe once
        # df_page_sheet = build_page_df(graph_client, service, start_date, end_date)
        # Fetch raw page sheet first
        df_page_sheet_raw = fetch_sheet_from_sharepoint(graph_client, "Page sheet")

        # Build final page dataframe with NEW columns appended
        df_page_sheet, rank_col, traffic_col = build_page_df(
            df_page_sheet_raw,
            df_gsc_page,
            end_date
        )
        
        logger.info("Page sheet data "+str(df_page_sheet.head()))

        logger.info(f"Page sheet shape: {df_page_sheet.shape}")

        # Normalize text fields used later
        df_query["query"] = df_query["query"].str.lower()
        df_gsc_page["page"] = df_gsc_page["page"].str.lower()
        df_query_page["query"] = df_query_page["query"].str.lower()
        df_query_page["page"] = df_query_page["page"].str.lower()

        for sheet_name in sheets_to_process:
            logger.info(f"Processing sheet: {sheet_name}")

            df_input = fetch_sheet_from_sharepoint(graph_client, "Sheet1")

            df_input.columns = (
                df_input.columns.astype(str)
                .str.strip()
                .str.lower()
                .str.replace(" ", "_")
            )

            df_input["blog_url"] = df_input["blog_url"].str.strip().str.lower()

            df_input["primary_keyword"] = (
                df_input["primary_keyword"].fillna('') + ' ' +
                df_input["secondary_keywords"].fillna('')
            ).str.strip().str.lower()

            summary_rows = []

            for _, row in df_input.iterrows():
                page = row["blog_url"]
                primary = row["primary_keyword"]

                secondary_raw = str(row.get("secondary_keywords", ""))
                secondary_list = [k.strip().lower() for k in secondary_raw.split(",") if k.strip()]
                all_keywords = [primary] + secondary_list

                matched = df_query_page[
                    (df_query_page["page"] == page) &
                    (df_query_page["query"].apply(lambda q: any(k in q for k in all_keywords)))
                ]

                if not matched.empty:
                    best = matched.sort_values(by="traffic", ascending=False).iloc[0]
                    summary_rows.append({
                        **row,
                        "Matched Query": best["query"],
                        "Traffic": best["traffic"],
                        "rank": best["rank"]
                    })
                else:
                    summary_rows.append({
                        **row,
                        "Matched Query": None,
                        "Traffic": None,
                        "rank": None
                    })

            df_summary = pd.DataFrame(summary_rows)

            output_file = f"output_{sheet_name}.xlsx"
            with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
                df_summary.to_excel(writer, sheet_name="Summary", index=False)
                df_query.to_excel(writer, sheet_name="Query_Data", index=False)
                df_gsc_page.to_excel(writer, sheet_name="Page_Data", index=False)
                df_country.to_excel(writer, sheet_name="Country_Data", index=False)
                df_device.to_excel(writer, sheet_name="Device_Data", index=False)
                df_query_page.to_excel(writer, sheet_name="Query_Page", index=False)

                df_page_sheet.to_excel(writer, sheet_name="Page sheet", index=False)
            logger.info(f"SEO report for {sheet_name} generated successfully!")

    except Exception as e:
        logger.exception(f"Pipeline failed: {str(e)}")        
def build_page_df(df_page_sheet, df_gsc_page, end_date):
    df_page_sheet = df_page_sheet.copy()
    df_gsc_page = df_gsc_page.copy()

    # ==========================================
    # 1) Flatten old grouped-header columns
    # Example:
    # urls | 2026-03-20 00:00:00 | Unnamed: 2
    # =>
    # urls | 3/20/2026_Rank | 3/20/2026_Traffic
    # ==========================================
    cols = list(df_page_sheet.columns)
    new_cols = []
    i = 0

    while i < len(cols):
        col = str(cols[i]).strip()

        # Keep URL column as-is
        if col.lower() == "urls":
            new_cols.append("urls")
            i += 1
            continue

        # Old grouped-header style:
        # first col = date, second col = Unnamed => treat as Rank + Traffic
        if "Unnamed:" not in col:
            next_col = str(cols[i + 1]).strip() if i + 1 < len(cols) else None

            if next_col and "Unnamed:" in next_col:
                # Convert date-like header to m/d/YYYY format
                try:
                    dt = pd.to_datetime(col)
                    date_label = f"{dt.month}/{dt.day}/{dt.year}"
                except Exception:
                    # fallback if parsing fails
                    date_label = col.split(" ")[0]

                new_cols.append(f"{date_label}_Rank")
                new_cols.append(f"{date_label}_Traffic")
                i += 2
                continue

        # Already flat column -> keep as-is
        new_cols.append(col)
        i += 1

    # Apply flattened column names
    if len(new_cols) == len(df_page_sheet.columns):
        df_page_sheet.columns = new_cols

    # ==========================================
    # 2) Create normalized temp copy for URL matching only
    # ==========================================
    df_temp = df_page_sheet.copy()
    df_temp.columns = (
        df_temp.columns.astype(str)
        .str.strip()
        .str.lower()
        .str.replace(" ", "_")
    )

    # Find URL column safely
    url_col = "urls"
    if url_col not in df_temp.columns:
        raise KeyError(f"Expected column '{url_col}' in Page sheet. Found: {list(df_temp.columns)}")

    # Normalize URLs in page sheet
    df_temp[url_col] = df_temp[url_col].astype(str).str.lower().str.strip()

    # Normalize GSC page URLs
    df_gsc_page["page"] = df_gsc_page["page"].astype(str).str.lower().str.strip()

    # ==========================================
    # 3) Build page-level metrics from GSC
    # ==========================================
    page_metrics = (
        df_gsc_page.groupby("page", as_index=False)
        .agg(
            traffic=("traffic", "sum"),
            rank=("rank", "min")
        )
    )

    metrics_lookup = page_metrics.set_index("page")

    # ==========================================
    # 4) New column names in SAME flat format
    # Example:
    # 3/30/2026_Rank
    # 3/30/2026_Traffic
    # ==========================================
    date_label = f"{end_date.month}/{end_date.day}/{end_date.year}"
    rank_col = f"{date_label}_Rank"
    traffic_col = f"{date_label}_Traffic"

    # Optional: overwrite if already exists (safe rerun)
    df_page_sheet[rank_col] = df_temp[url_col].map(metrics_lookup["rank"])
    df_page_sheet[traffic_col] = df_temp[url_col].map(metrics_lookup["traffic"])

    return df_page_sheet, rank_col, traffic_col        
def write_page_sheet_with_all_groups(
    writer,
    df,
    sheet_name="Page sheet",
    base_column="urls",
    base_column_width=90,
    metric_column_width=18
):
    """
    Writes Page sheet in grouped header format:

    Row 1: urls | 2026-03-20 | 2026-03-30 | ...
    Row 2:      | Rank | Traffic | Rank | Traffic | ...
    Row 3+: data

    Supports:
    - old flattened columns from previous grouped headers
      e.g. '2026-03-20 00:00:00' + 'Unnamed: 2'
    - new flat columns like '2026-03-30_rank', '2026-03-30_traffic'
    """

    workbook = writer.book

    # Remove existing sheet if already exists
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    ws = workbook.create_sheet(title=sheet_name)

    # -------------------------
    # Detect metric column pairs
    # -------------------------
    cols = list(df.columns)

    # Must contain base column
    if base_column not in cols:
        raise KeyError(f"'{base_column}' column not found in df columns: {cols}")

    grouped_metrics = []  # [(date_label, rank_col, traffic_col)]

    i = 0
    while i < len(cols):
        col = str(cols[i])

        if col == base_column:
            i += 1
            continue

        # CASE 1: New style columns => YYYY-MM-DD_rank + YYYY-MM-DD_traffic
        if col.endswith("_rank"):
            date_label = col[:-5]  # remove "_rank"
            traffic_col = f"{date_label}_traffic"
            if traffic_col in cols:
                grouped_metrics.append((date_label, col, traffic_col))
                i += 1
                continue

        # CASE 2: Old style flattened grouped header
        # e.g. '2026-03-20 00:00:00' followed by 'Unnamed: 2'
        if "Unnamed:" not in col:
            next_col = str(cols[i + 1]) if i + 1 < len(cols) else None
            if next_col and "Unnamed:" in next_col:
                # normalize date label
                date_label = col.split(" ")[0]
                grouped_metrics.append((date_label, col, next_col))
                i += 2
                continue

        i += 1

    # -------------------------
    # Write Row 1
    # -------------------------
    ws.cell(row=1, column=1, value=base_column)

    current_col = 2
    for date_label, rank_col, traffic_col in grouped_metrics:
        ws.cell(row=1, column=current_col, value=date_label)
        ws.merge_cells(
            start_row=1,
            start_column=current_col,
            end_row=1,
            end_column=current_col + 1
        )
        current_col += 2

    # -------------------------
    # Write Row 2
    # -------------------------
    ws.cell(row=2, column=1, value="")

    current_col = 2
    for _, _, _ in grouped_metrics:
        ws.cell(row=2, column=current_col, value="Rank")
        ws.cell(row=2, column=current_col + 1, value="Traffic")
        current_col += 2

    # -------------------------
    # Write Data
    # -------------------------
    for row_idx, (_, row) in enumerate(df.iterrows(), start=3):
        ws.cell(row=row_idx, column=1, value=row[base_column])

        current_col = 2
        for _, rank_col, traffic_col in grouped_metrics:
            ws.cell(row=row_idx, column=current_col, value=row.get(rank_col))
            ws.cell(row=row_idx, column=current_col + 1, value=row.get(traffic_col))
            current_col += 2

    # -------------------------
    # Column widths
    # -------------------------
    ws.column_dimensions["A"].width = base_column_width

    for col_idx in range(2, current_col):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = metric_column_width        
        
def fetch_sheet_from_sharepoint(client, sheet_name):
    logger.info(f"Fetching sheet '{sheet_name}' from SharePoint...")

    file_url = f"https://graph.microsoft.com/v1.0/sites/{SITE_ID}/drives/{DRIVE_ID}/items/{FILE_ID}/content"
    response = requests.get(file_url, headers=client.get_headers())

    if response.status_code != 200:
        logger.error(f"SharePoint Fetch Error: {response.text}")
        raise Exception(f"SharePoint Fetch Error: {response.text}")

    xls = pd.ExcelFile(io.BytesIO(response.content))

    # Read first row as header, skip second row, start data from third row
    df = pd.read_excel(
        xls,
        sheet_name=sheet_name,
        header=0,
    )

    logger.info(f"Loaded sheet '{sheet_name}' with shape: {df.shape}")
    return df
# =========================
# ▶️ RUN
# =========================
if __name__ == "__main__":
    main()